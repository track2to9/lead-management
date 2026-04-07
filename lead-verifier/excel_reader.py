"""
엑셀 파일에서 업체 리스트 파싱
GPT가 생성한 비정형 엑셀 → 구조화된 딕셔너리 리스트로 변환
"""
import re
from openpyxl import load_workbook


def parse_dealer_list(filepath: str) -> list[dict]:
    """GPT가 만든 딜러 리스트 엑셀을 파싱"""
    wb = load_workbook(filepath)
    ws = wb.active

    companies = []
    current = None
    current_grade = ""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        val = str(row[0]).strip() if row[0] else ""
        val2 = str(row[1]).strip() if row[1] else ""
        combined = (val + " " + val2).strip()

        if not combined:
            continue

        # 섹션 헤더 (A등급, B등급)
        if "A. 확인 강도" in combined:
            current_grade = "A"
            continue
        elif "B. " in combined and ("확장" in combined or "후보" in combined):
            current_grade = "B"
            continue
        elif "C. " in combined:
            current_grade = "C"
            continue

        # 업체명 (숫자. 으로 시작)
        match = re.match(r"(\d+)\.\s+(.+)", combined)
        if match:
            if current:
                companies.append(current)
            current = {
                "num": int(match.group(1)),
                "name": match.group(2).strip(),
                "grade": current_grade,
                "url": "",
                "emails_gpt": [],
                "products": "",
                "memo": "",
                # 검증 결과 (크롤링 후 채움)
                "url_accessible": None,
                "emails_found": [],
                "phone_found": [],
                "contact_page": "",
                "verification_status": "pending",
                "screenshot_path": "",
            }
            continue

        if current is None:
            continue

        # 필드 파싱
        if "홈페이지:" in combined:
            urls = re.findall(r"https?://[^\s,]+", combined)
            current["url"] = urls[0] if urls else ""
        elif "이메일:" in combined:
            emails = re.findall(r"[\w.+-]+@[\w-]+\.[\w.]+", combined)
            current["emails_gpt"] = emails
        elif "취급:" in combined:
            current["products"] = combined.replace("취급:", "").strip()
        elif "메모:" in combined:
            current["memo"] = combined.replace("메모:", "").strip()

    if current:
        companies.append(current)

    return companies


if __name__ == "__main__":
    filepath = "/Users/youngminkim/Sites/ai.devpartner.org/case-study/개인사업자-건설부품 제공/폴란드 굴삭기 어태치 딜러 리스트.xlsx"
    companies = parse_dealer_list(filepath)

    print(f"총 {len(companies)}개 업체 파싱 완료\n")

    for c in companies:
        email_status = "✓" if c["emails_gpt"] else "✗"
        url_status = "✓" if c["url"] else "✗"
        print(
            f"  [{c['grade']}] {c['num']:>2}. {c['name'][:40]:<40} "
            f"URL:{url_status} Email:{email_status} ({len(c['emails_gpt'])}개)"
        )
