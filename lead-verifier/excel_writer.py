"""
검증 결과를 엑셀 파일로 출력
고객에게 전달할 깔끔한 형태
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def write_results(companies: list[dict], output_path: str):
    """검증 결과를 엑셀로 출력"""
    wb = Workbook()
    ws = wb.active
    ws.title = "검증 결과"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 헤더
    headers = [
        "No", "등급", "매칭점수", "우선순위", "업체명", "홈페이지", "접속여부",
        "GPT 이메일", "크롤링 이메일", "신규 발견 이메일",
        "전화번호", "Contact 페이지", "취급 제품",
        "회사 요약", "SPS 매칭 이유", "접근 전략",
        "검증 상태", "비고"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # 데이터
    high_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    medium_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for row_idx, c in enumerate(companies, 2):
        v = c.get("validation", {})
        a = c.get("analysis", {})

        data = [
            c["num"],
            c["grade"],
            a.get("match_score", 0),
            a.get("priority", ""),
            c["name"],
            c["url"],
            "O" if c.get("url_accessible") else "X",
            "\n".join(c.get("emails_gpt", [])),
            "\n".join(c.get("emails_found", [])),
            "\n".join(v.get("emails_new", [])),
            "\n".join(c.get("phone_found", [])[:3]),
            c.get("contact_page", ""),
            c.get("products", "")[:80],
            a.get("summary", ""),
            a.get("match_reason", ""),
            a.get("approach", ""),
            _status_label(v.get("overall", "")),
            c.get("memo", "")[:60],
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

        # 우선순위별 행 색상
        priority = a.get("priority", "")
        if priority == "high":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = high_fill
        elif priority == "medium":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = medium_fill

        # 검증 상태 색상
        status = v.get("overall", "")
        status_col = len(headers) - 1  # 검증 상태 컬럼
        status_cell = ws.cell(row=row_idx, column=status_col)
        if status == "verified":
            status_cell.fill = verified_fill
        elif status == "url_fail":
            status_cell.fill = error_fill
        elif status == "no_email_found":
            status_cell.fill = warning_fill

    # 컬럼 너비
    col_letters = "ABCDEFGHIJKLMNOPQR"
    widths = [5, 5, 8, 8, 30, 35, 8, 28, 28, 28, 18, 30, 35, 40, 40, 40, 12, 30]
    for i, w in enumerate(widths):
        if i < len(col_letters):
            ws.column_dimensions[col_letters[i]].width = w

    # 필터
    ws.auto_filter.ref = ws.dimensions

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    total = len(companies)
    verified = sum(1 for c in companies if c.get("validation", {}).get("overall") == "verified")
    updated = sum(1 for c in companies if c.get("validation", {}).get("overall") == "updated")
    total_new = sum(len(c.get("validation", {}).get("emails_new", [])) for c in companies)

    summary_data = [
        ["항목", "값"],
        ["검증 일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["총 업체 수", total],
        ["검증 완료 (이메일 일치)", verified],
        ["정보 업데이트 (신규 이메일 발견)", updated],
        ["신규 발견 이메일 수", total_new],
        ["", ""],
        ["생성", "AI DevPartner (ai.devpartner.org)"],
    ]

    for r, row_data in enumerate(summary_data, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 40

    wb.save(output_path)
    print(f"\n📄 검증 결과 저장: {output_path}")


def _status_label(status: str) -> str:
    labels = {
        "verified": "✅ 확인됨",
        "updated": "🔄 업데이트됨",
        "url_fail": "❌ 접속실패",
        "no_email_found": "⚠️ 이메일없음",
        "no_url": "- URL없음",
        "pending": "⏳ 대기",
        "unknown": "? 미확인",
    }
    return labels.get(status, status)
