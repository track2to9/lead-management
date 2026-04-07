"""
UAngel 5G Core NF - 일본 Local 5G 잠재 파트너 분석
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

입력: 일본 Local 5G SI/벤더 리스트 (엑셀)
출력: 홈페이지 분석 + 매칭 점수 + 접근 전략 (엑셀)

실행: python main_japan.py
"""
import asyncio
import sys
import os
from datetime import datetime
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(__file__))

from web_scraper_japan import scrape_all
from validator import validate_all, print_summary
from excel_writer import write_results
from company_analyzer_japan import analyze_company


def parse_structured_excel(filepath: str) -> list[dict]:
    """구조화된 엑셀 파일 파싱 (헤더 행 포함)"""
    wb = load_workbook(filepath)
    ws = wb.active

    companies = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue
        companies.append({
            "num": int(row[0]),
            "grade": str(row[1] or "B"),
            "name": str(row[2] or ""),
            "url": str(row[3] or ""),
            "emails_gpt": [e.strip() for e in str(row[4] or "").split(",") if "@" in str(row[4] or "")],
            "products": str(row[5] or ""),
            "memo": str(row[6] or ""),
            "url_accessible": None,
            "emails_found": [],
            "phone_found": [],
            "contact_page": "",
            "verification_status": "pending",
            "screenshot_path": "",
        })

    return companies


async def main():
    # ── 설정 ──
    BASE_DIR = os.path.dirname(__file__)
    INPUT_FILE = os.path.join(BASE_DIR, "input/uangel_japan_local5g_buyers.xlsx")
    OUTPUT_DIR = os.path.join(BASE_DIR, "output")
    SCREENSHOT_DIR = os.path.join(BASE_DIR, "screenshots_japan")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(SCREENSHOT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"UAngel_일본Local5G_분석결과_{timestamp}.xlsx")

    # ── Step 1: 엑셀 파싱 ──
    print("=" * 60)
    print("  UAngel 5G Core NF - 일본 Local 5G 파트너 분석")
    print("  TradeVoy (trade.devpartner.org)")
    print("=" * 60)
    print(f"\n📂 입력 파일: {INPUT_FILE}")

    companies = parse_structured_excel(INPUT_FILE)
    print(f"✅ {len(companies)}개 업체 파싱 완료")

    # ── Step 2: 홈페이지 크롤링 ──
    print(f"\n🌐 홈페이지 크롤링 시작 (일본어 사이트)...")
    companies = await scrape_all(companies, SCREENSHOT_DIR)

    # ── Step 3: 검증 ──
    print(f"\n🔍 GPT 데이터 vs 실제 데이터 검증 중...")
    companies = validate_all(companies)

    # ── Step 4: Claude AI 분석 (UAngel 매칭) ──
    print(f"\n🤖 Claude AI 분석 중 (UAngel 5G Core NF 매칭)...")
    for c in companies:
        page_text = c.get("page_text", "")
        if c.get("url_accessible"):
            await analyze_company(c, page_text)
        else:
            c["analysis"] = {
                "summary": "홈페이지 접속 불가",
                "match_score": 0,
                "match_reason": "분석 불가",
                "approach": "",
                "priority": "low",
            }

    # 매칭 점수 순으로 정렬
    companies.sort(key=lambda x: x.get("analysis", {}).get("match_score", 0), reverse=True)

    # ── Step 5: 결과 출력 ──
    print_summary(companies)

    print(f"\n{'='*60}")
    print(f"  UAngel 매칭 순위 (상위)")
    print(f"{'='*60}")
    for c in companies[:15]:
        a = c.get("analysis", {})
        print(f"  {a.get('match_score', 0):>3}점 [{a.get('priority', '?'):>6}] {c['name'][:40]}")
        print(f"       → {a.get('match_reason', '')[:80]}")
    print(f"{'='*60}")

    write_results(companies, OUTPUT_FILE)

    print(f"\n📸 스크린샷: {SCREENSHOT_DIR}/")
    print(f"📄 결과 파일: {OUTPUT_FILE}")
    print(f"\n✅ 완료!")


if __name__ == "__main__":
    asyncio.run(main())
