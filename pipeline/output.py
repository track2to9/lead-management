"""
결과 출력 모듈
- CSV 출력 (UTF-8 BOM, Excel 한글 호환)
- Excel 출력 (lead-verifier/excel_writer.py 패턴 재사용)
- Zoho CRM 호환 컬럼 포맷
- 요약 통계

TODO (Supabase persistence): Evidence items collected in candidate["evidence"] are
currently included in the CSV/Excel output as a summary string in the Notes column.
They are NOT yet written to the Supabase `evidence` table. When the pipeline gains
Supabase integration, add an upsert loop here (or in batch_runner) to persist each
evidence item with its candidate FK, source_url, source_type, screenshot_path,
text_excerpt, text_translated, related_scores, and content_date.
"""
import csv
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ─── 헬퍼 ────────────────────────────────────────────────────

def _build_notes(company: dict) -> str:
    """
    Build a human-readable notes string for CSV/Excel output.

    Combines any pipeline warning with a summary of collected evidence items.
    Evidence is not yet persisted to Supabase — see the module-level TODO above.
    """
    parts: list[str] = []

    warning = company.get("_warning", "")
    if warning:
        parts.append(warning)

    evidence: list[dict] = company.get("evidence") or []
    if evidence:
        source_types = [item.get("source_type", "") for item in evidence if item.get("source_type")]
        types_str = ", ".join(source_types) if source_types else "unknown"
        parts.append(f"[evidence: {len(evidence)}건 / {types_str}]")

    return " | ".join(parts)


# ─── CSV 출력 ───────────────────────────────────────────────

def write_csv(results: list[dict], output_path: str, config_summary: dict | None = None):
    """
    CSV 파일 출력 (UTF-8 BOM - Excel 한글 호환).
    Zoho CRM 호환 컬럼 포맷.

    Args:
        results: run_batch()의 반환값 (국가별 결과 리스트)
        output_path: 출력 파일 경로
        config_summary: 설정 요약 정보 (상단 코멘트 출력용)
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # 모든 국가의 업체를 하나로 병합
    all_companies = []
    for r in results:
        country = r.get("country", "")
        for c in r.get("companies", []):
            c["_country"] = country
            all_companies.append(c)

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)

        # 상단 요약 (코멘트 행)
        if config_summary:
            writer.writerow([f"# 생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
            writer.writerow([f"# 클라이언트: {config_summary.get('client', '')}"])
            writer.writerow([f"# 제품: {config_summary.get('product', '')}"])
            writer.writerow([f"# 국가: {', '.join(config_summary.get('countries', []))}"])
            writer.writerow([f"# 총 업체 수: {len(all_companies)}"])
            writer.writerow([])

        # Zoho CRM 호환 헤더 (인텔리전스 + 팔로업 + 전시회 확장)
        headers = [
            "Company",          # 회사명
            "Website",          # 웹사이트
            "Country",          # 국가
            "Email",            # 대표 이메일
            "Phone",            # 전화번호
            "Source",           # 데이터 출처
            "Match Score",      # 매칭 점수
            "Product Fit Score",
            "Buying Signal Score",
            "Company Capability Score",
            "Accessibility Score",
            "Strategic Value Score",
            "Priority",         # 우선순위
            "Match Reason",     # 매칭 사유
            "Approach Strategy", # 접근 전략
            "Company Summary",  # 회사 요약
            "Products",         # 취급 제품
            # 바이어 인텔리전스 확장
            "Current Suppliers",     # 현재 공급업체
            "Company Size",          # 회사 규모 추정
            "Decision Maker",        # 의사결정권자 추정
            "Best Timing",           # 최적 접근 시기
            "Competitive Landscape", # 경쟁 환경
            # 기존 필드
            "Contact Page",     # 연락처 페이지
            "Verification",     # 검증 상태
            "Email Subject",    # 이메일 제목 (초안)
            "Email Body",       # 이메일 본문 (초안)
            # 팔로업 시퀀스
            "Followup Day3 Subject",
            "Followup Day3 Body",
            "Followup Day7 Subject",
            "Followup Day7 Body",
            "Followup Day14 Subject",
            "Followup Day14 Body",
            # 전시회
            "Matched Exhibitions",
            # 메타
            "Data Source Type", # 소스 유형 (directory/llm)
            "Notes",            # 비고
        ]
        writer.writerow(headers)

        for c in all_companies:
            analysis = c.get("analysis", {})
            validation = c.get("validation", {})
            email_draft = c.get("email_draft", {})
            emails = c.get("emails_found", [])
            phones = c.get("phone_found", [])
            followups = c.get("followup_sequence", {}).get("emails", [])
            exhibitions = c.get("matched_exhibitions", [])

            # 검증 상태 라벨
            status = validation.get("overall", c.get("verification_status", "UNVERIFIED"))

            # 팔로업 이메일 (day별 매핑)
            fu = {e.get("day"): e for e in followups}

            row = [
                c.get("name", ""),
                c.get("url", ""),
                c.get("_country", ""),
                "; ".join(emails) if emails else "",
                "; ".join(phones[:3]) if phones else "",
                c.get("source", ""),
                analysis.get("match_score", 0),
                *(
                    lambda bd: [
                        bd.get("product_fit", {}).get("score", ""),
                        bd.get("buying_signal", {}).get("score", ""),
                        bd.get("company_capability", {}).get("score", ""),
                        bd.get("accessibility", {}).get("score", ""),
                        bd.get("strategic_value", {}).get("score", ""),
                    ]
                )(analysis.get("score_breakdown", {})),
                analysis.get("priority", ""),
                analysis.get("match_reason", ""),
                analysis.get("approach", ""),
                analysis.get("summary", ""),
                ", ".join(analysis.get("detected_products", [])),
                # 인텔리전스
                ", ".join(analysis.get("current_suppliers", [])),
                analysis.get("company_size_estimate", ""),
                analysis.get("decision_maker_hint", ""),
                analysis.get("best_timing", ""),
                analysis.get("competitive_landscape", ""),
                # 기존
                c.get("contact_page", ""),
                status,
                email_draft.get("subject", ""),
                email_draft.get("body", ""),
                # 팔로업
                fu.get(3, {}).get("subject", ""),
                fu.get(3, {}).get("body", ""),
                fu.get(7, {}).get("subject", ""),
                fu.get(7, {}).get("body", ""),
                fu.get(14, {}).get("subject", ""),
                fu.get(14, {}).get("body", ""),
                # 전시회
                "; ".join(e.get("name", "") for e in exhibitions[:3]),
                # 메타
                c.get("source_type", "unknown"),
                _build_notes(c),
            ]
            writer.writerow(row)

    print(f"\n  📄 CSV 저장: {output_path}")
    print(f"     (UTF-8 BOM - Excel에서 한글 깨짐 없이 열림)")


# ─── Excel 출력 ─────────────────────────────────────────────

def write_excel(results: list[dict], output_path: str, config_summary: dict | None = None):
    """
    Excel 파일 출력 (lead-verifier/excel_writer.py 패턴 재사용).
    우선순위별 색상, 필터, 요약 시트 포함.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()

    # 모든 국가 병합
    all_companies = []
    for r in results:
        country = r.get("country", "")
        for c in r.get("companies", []):
            c["_country"] = country
            all_companies.append(c)

    # ── 메인 시트: 잠재고객 리스트 ──
    ws = wb.active
    ws.title = "잠재고객 리스트"

    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    high_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    medium_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "No", "국가", "매칭점수", "제품적합도", "구매시그널", "기업역량", "접근성", "전략가치", "우선순위", "업체명", "홈페이지",
        "이메일", "전화번호", "데이터 출처", "소스 유형",
        "회사 요약", "매칭 사유", "접근 전략", "취급 제품",
        "현재 공급업체", "회사 규모", "의사결정권자", "접근 타이밍", "경쟁 환경",
        "검증 상태", "이메일 제목(초안)", "비고",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, c in enumerate(all_companies, 2):
        analysis = c.get("analysis", {})
        validation = c.get("validation", {})
        email_draft = c.get("email_draft", {})

        status = validation.get("overall", c.get("verification_status", "UNVERIFIED"))

        _bd = analysis.get("score_breakdown", {})
        data = [
            row_idx - 1,
            c.get("_country", ""),
            analysis.get("match_score", 0),
            _bd.get("product_fit", {}).get("score", ""),
            _bd.get("buying_signal", {}).get("score", ""),
            _bd.get("company_capability", {}).get("score", ""),
            _bd.get("accessibility", {}).get("score", ""),
            _bd.get("strategic_value", {}).get("score", ""),
            analysis.get("priority", ""),
            c.get("name", ""),
            c.get("url", ""),
            "; ".join(c.get("emails_found", [])),
            "; ".join(c.get("phone_found", [])[:3]),
            c.get("source", ""),
            c.get("source_type", ""),
            analysis.get("summary", ""),
            analysis.get("match_reason", ""),
            analysis.get("approach", ""),
            ", ".join(analysis.get("detected_products", [])),
            ", ".join(analysis.get("current_suppliers", [])),
            analysis.get("company_size_estimate", ""),
            analysis.get("decision_maker_hint", ""),
            analysis.get("best_timing", ""),
            analysis.get("competitive_landscape", ""),
            status,
            email_draft.get("subject", ""),
            _build_notes(c),
        ]

        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

        # 우선순위별 행 색상
        priority = analysis.get("priority", "")
        if priority == "high":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = high_fill
        elif priority == "medium":
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = medium_fill

        # 검증 상태 색상 (25번째 컬럼: 인텔리전스 5개 + 점수 세분화 5개 추가됨)
        status_col = 25
        status_cell = ws.cell(row=row_idx, column=status_col)
        if status == "verified":
            status_cell.fill = verified_fill
        elif status in ("url_fail", "error"):
            status_cell.fill = error_fill
        elif status in ("no_email_found", "UNVERIFIED"):
            status_cell.fill = warning_fill

    # 컬럼 너비 (인텔리전스 5개 + 점수 세분화 5개 컬럼 추가)
    widths = [5, 10, 8, 8, 8, 8, 8, 8, 8, 30, 35, 30, 18, 20, 12, 40, 40, 40, 30, 25, 12, 20, 25, 30, 12, 35, 20]
    for i, w in enumerate(widths):
        col_letter = chr(65 + i) if i < 26 else chr(64 + i // 26) + chr(65 + i % 26)
        ws.column_dimensions[col_letter].width = w

    ws.auto_filter.ref = ws.dimensions

    # ── 요약 시트 ──
    ws2 = wb.create_sheet("요약")

    summary_data = [
        ["항목", "값"],
        ["생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["", ""],
    ]

    if config_summary:
        summary_data.extend([
            ["클라이언트", config_summary.get("client", "")],
            ["제품 카테고리", config_summary.get("product", "")],
            ["대상 국가", ", ".join(config_summary.get("countries", []))],
            ["LLM 프로바이더", config_summary.get("llm_provider", "")],
            ["", ""],
        ])

    # 국가별 통계
    for r in results:
        stats = r.get("stats", {})
        country = r.get("country", "?")
        if r.get("error"):
            summary_data.append([f"{country} (실패)", r["error"]])
            continue
        summary_data.extend([
            [f"── {country} ──", ""],
            ["  총 업체", stats.get("total", 0)],
            ["  디렉토리 수집", stats.get("from_directories", 0)],
            ["  LLM 생성", stats.get("from_llm", 0)],
            ["  URL 접근 가능", stats.get("url_accessible", 0)],
            ["  High 매칭", stats.get("high_match", 0)],
            ["  Medium 매칭", stats.get("medium_match", 0)],
            ["  평균 매칭 점수", stats.get("avg_match_score", 0)],
            ["  발견 이메일 수", stats.get("emails_found", 0)],
            ["  이메일 초안 수", stats.get("emails_drafted", 0)],
            ["  할루시네이션 비율", f"{stats.get('hallucination_rate_pct', 0)}%"],
            ["", ""],
        ])

    summary_data.extend([
        ["", ""],
        ["생성", "AI DevPartner Pipeline (ai.devpartner.org)"],
    ])

    for r, row_data in enumerate(summary_data, 1):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c_idx, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 45

    # ── 이메일 초안 시트 ──
    ws3 = wb.create_sheet("이메일 초안")

    email_headers = ["No", "국가", "업체명", "매칭점수", "이메일주소", "제목", "본문"]
    for col, h in enumerate(email_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    email_row = 2
    for c in all_companies:
        draft = c.get("email_draft", {})
        if not draft.get("subject"):
            continue
        data = [
            email_row - 1,
            c.get("_country", ""),
            c.get("name", ""),
            c.get("analysis", {}).get("match_score", 0),
            "; ".join(c.get("emails_found", [])),
            draft.get("subject", ""),
            draft.get("body", ""),
        ]
        for col, val in enumerate(data, 1):
            cell = ws3.cell(row=email_row, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
        email_row += 1

    ws3.column_dimensions["A"].width = 5
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["D"].width = 8
    ws3.column_dimensions["E"].width = 30
    ws3.column_dimensions["F"].width = 40
    ws3.column_dimensions["G"].width = 80

    # ── 팔로업 시퀀스 시트 ──
    ws4 = wb.create_sheet("팔로업 시퀀스")

    fu_headers = ["No", "국가", "업체명", "매칭점수", "단계", "발송일", "제목", "본문"]
    for col, h in enumerate(fu_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    fu_row = 2
    fu_num = 1
    for c in all_companies:
        followups = c.get("followup_sequence", {}).get("emails", [])
        if not followups:
            continue
        for fu in followups:
            data = [
                fu_num,
                c.get("_country", ""),
                c.get("name", ""),
                c.get("analysis", {}).get("match_score", 0),
                fu.get("label", f"Day {fu.get('day', '?')}"),
                f"+{fu.get('day', '?')}일",
                fu.get("subject", ""),
                fu.get("body", ""),
            ]
            for col, val in enumerate(data, 1):
                cell = ws4.cell(row=fu_row, column=col, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
            fu_row += 1
        fu_num += 1

    ws4.column_dimensions["A"].width = 5
    ws4.column_dimensions["B"].width = 10
    ws4.column_dimensions["C"].width = 25
    ws4.column_dimensions["D"].width = 8
    ws4.column_dimensions["E"].width = 10
    ws4.column_dimensions["F"].width = 8
    ws4.column_dimensions["G"].width = 40
    ws4.column_dimensions["H"].width = 80

    # ── 바이어 인텔리전스 시트 ──
    ws5 = wb.create_sheet("바이어 인텔리전스")

    intel_headers = [
        "No", "국가", "업체명", "매칭점수", "현재 공급업체",
        "회사 규모", "의사결정권자", "접근 타이밍", "경쟁 환경", "관련 전시회",
    ]
    for col, h in enumerate(intel_headers, 1):
        cell = ws5.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    intel_row = 2
    for idx, c in enumerate(all_companies, 1):
        analysis = c.get("analysis", {})
        if analysis.get("match_score", 0) < 30:
            continue
        exhibitions = c.get("matched_exhibitions", [])
        data = [
            idx,
            c.get("_country", ""),
            c.get("name", ""),
            analysis.get("match_score", 0),
            ", ".join(analysis.get("current_suppliers", [])),
            analysis.get("company_size_estimate", ""),
            analysis.get("decision_maker_hint", ""),
            analysis.get("best_timing", ""),
            analysis.get("competitive_landscape", ""),
            "; ".join(e.get("name", "") for e in exhibitions[:3]),
        ]
        for col, val in enumerate(data, 1):
            cell = ws5.cell(row=intel_row, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border
        intel_row += 1

    ws5.column_dimensions["A"].width = 5
    ws5.column_dimensions["B"].width = 10
    ws5.column_dimensions["C"].width = 25
    ws5.column_dimensions["D"].width = 8
    ws5.column_dimensions["E"].width = 30
    ws5.column_dimensions["F"].width = 15
    ws5.column_dimensions["G"].width = 20
    ws5.column_dimensions["H"].width = 30
    ws5.column_dimensions["I"].width = 40
    ws5.column_dimensions["J"].width = 35

    # ── 전시회 시트 ──
    ws6 = wb.create_sheet("관련 전시회")

    ex_headers = [
        "No", "전시회명", "개최 장소", "시기", "빈도",
        "규모", "관련성", "대상 방문자", "웹사이트", "추천 액션",
    ]
    for col, h in enumerate(ex_headers, 1):
        cell = ws6.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # 전시회 데이터는 results 안에 exhibitions 키로 있을 수 있음
    ex_row = 2
    seen_exhibitions = set()
    for r in results:
        for ex in r.get("exhibitions", []):
            ex_name = ex.get("name", "")
            if ex_name in seen_exhibitions:
                continue
            seen_exhibitions.add(ex_name)
            data = [
                ex_row - 1,
                ex_name,
                ex.get("location", ""),
                ex.get("typical_month", ""),
                ex.get("frequency", ""),
                ex.get("estimated_size", ""),
                ex.get("relevance", ""),
                ex.get("target_visitors", ""),
                ex.get("website", ""),
                ex.get("action_suggestion", ""),
            ]
            for col, val in enumerate(data, 1):
                cell = ws6.cell(row=ex_row, column=col, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
            ex_row += 1

    ws6.column_dimensions["A"].width = 5
    ws6.column_dimensions["B"].width = 30
    ws6.column_dimensions["C"].width = 20
    ws6.column_dimensions["D"].width = 15
    ws6.column_dimensions["E"].width = 10
    ws6.column_dimensions["F"].width = 15
    ws6.column_dimensions["G"].width = 35
    ws6.column_dimensions["H"].width = 25
    ws6.column_dimensions["I"].width = 35
    ws6.column_dimensions["J"].width = 35

    wb.save(output_path)
    print(f"  📊 Excel 저장: {output_path}")
